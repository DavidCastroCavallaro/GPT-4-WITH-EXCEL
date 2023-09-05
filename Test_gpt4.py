import gpt4
import openpyxl

def get_data_from_excel():
  """Gets the data from the specified worksheet in the specified Excel spreadsheet.

  Prompts the user for the excel_file and worksheet_name.

  Returns:
    A list of lists, where each inner list contains the data from one row in the worksheet.
  """

  excel_file = input("Enter the name of the Excel spreadsheet: ")
  worksheet_name = input("Enter the name of the worksheet to be processed: ")

  with open(excel_file, "rb") as f:
    wb = openpyxl.load_workbook(f)
    ws = wb[worksheet_name]

    data = []
    for row in ws.iter_rows():
      data.append([cell.value for cell in row])

    return data

def sort_data(data):
  """Sorts the data in ascending order.

  Args:
    data: The data to be sorted.

  Returns:
    A sorted list of the data.
  """

  sorted_data = gpt4.generate(
      prompt="Write code to sort the data in ascending order.",
      temperature=0.7,
      max_tokens=100
  )

  return sorted_data

def main():
  data = get_data_from_excel()

  # Ask the user how GPT should understand and organize the data in a new worksheet.
  organization_method = input(
      "How should GPT understand and organize the data in a new worksheet? "
  )

  sorted_data = sort_data(data)

  with open(excel_file, "rb") as f:
    wb = openpyxl.load_workbook(f)
    ws = wb.create_sheet("Sorted_Data")

    for row in sorted_data:
      ws.append(row)

    # Apply the organization method specified by the user using GPT.
    ws.auto_filter.filter_column(gpt4.generate(
      prompt="Write code to filter the data based on the organization method.",
      temperature=0.7,
      max_tokens=100
    ))

    wb.save(excel_file)

if __name__ == "__main__":
  main()
